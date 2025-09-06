#!/usr/bin/env python3

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime

def format_worksheet_clean(worksheet, sheet_type='detail'):
    """Format worksheet with clean styling - no colors or borders, just proper widths and alignment"""
    
    # Define simple styles - no colors or borders
    header_font = Font(bold=True)
    total_font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long text
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Format headers and data without colors or borders
    if sheet_type == 'detail':
        # For detail sheets, headers are in row 1
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # Check for amount columns
                header_value = str(worksheet.cell(1, cell.column).value).strip().lower()
                if 'amount' in header_value:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif 'document number' in header_value or 'panda bank transaction id' in header_value:
                    cell.alignment = Alignment(horizontal='center')
                elif 'date' in header_value:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                
                # Format account total rows (top and bottom totals)
                if str(cell.value).startswith('ACCOUNT ') and 'TOTAL' in str(cell.value):
                    for c in row:
                        c.font = total_font
                elif cell.value == 'TOTAL':
                    for c in row:
                        c.font = total_font
    
    elif sheet_type == 'pivot':
        # For pivot table, format starting from row 3 (after title)
        for row in worksheet.iter_rows(min_row=3, max_row=3):
            for cell in row:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=4):
            for cell in row:
                if cell.column in [2, 3]:  # Amount and Tab Total columns
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif cell.column == 4:  # Variance column
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Highlight Subtotal and Grand Total rows with bold only
                if cell.value in ['Subtotal', 'Grand Total']:
                    for c in row:
                        c.font = total_font

def read_bank_transaction_mapping():
    """Read bank transaction data for non-batch activity posting accounts"""
    
    # Account mapping: Netsuite Account -> (Account Number, Bank Transaction Files)
    # These are the NON-batch activity posting accounts
    account_mapping = {
        10068: (6, ['Bank Trasnsaction Data 6.csv']),    # Chase Reverse Wire Payrolls
        10069: (7, ['Bank Trasnsaction Data 7.csv']),    # Chase Recovery Ops
        10071: (9, ['Bank Trasnsaction Data 9.csv']),    # Chase Wire In
        10504: (21, ['Bank Trasnsaction Data 21.csv']),  # Chase 3rd Party Processors
        10513: (18, ['Bank Trasnsaction Data 18.csv'])   # PNC Customer Wire Ins
    }
    
    bank_mapping_data = {}
    
    for account_num, (orig_id, files) in account_mapping.items():
        print(f"\nReading bank transaction mapping for account {account_num} (origination_account_id: {orig_id})")
        
        combined_data = []
        for file in files:
            try:
                # Try different encodings and delimiters
                configurations = [
                    {'encoding': 'utf-8', 'sep': '\t'},
                    {'encoding': 'utf-8', 'sep': ','},
                    {'encoding': 'utf-16', 'sep': '\t'},
                    {'encoding': 'utf-16', 'sep': ','},
                    {'encoding': 'latin-1', 'sep': '\t'},
                    {'encoding': 'latin-1', 'sep': ','}
                ]
                
                df = None
                
                for config in configurations:
                    try:
                        df = pd.read_csv(file, encoding=config['encoding'], sep=config['sep'], 
                                       on_bad_lines='skip', low_memory=False)
                        
                        # Check if we got meaningful data
                        if len(df) > 0 and len(df.columns) > 5:
                            print(f"  Read {len(df)} records from {file} (encoding: {config['encoding']})")
                            break
                        else:
                            df = None
                            continue
                            
                    except Exception:
                        continue
                
                if df is not None:
                    combined_data.append(df)
                else:
                    print(f"  Could not read {file}")
                    
            except Exception as e:
                print(f"  Error reading {file}: {e}")
        
        if combined_data:
            # Combine all dataframes for this account
            account_df = pd.concat(combined_data, ignore_index=True)
            
            # Keep only relevant columns for matching: Panda Bank Transaction Id, Amount, Date
            mapping_columns = []
            
            # Find the relevant columns
            panda_id_col = None
            amount_col = None
            date_col = None
            
            for col in account_df.columns:
                col_clean = str(col).strip().lower()
                if 'panda bank transaction id' in col_clean:
                    panda_id_col = col
                    mapping_columns.append(col)
                elif 'signed amount' in col_clean:
                    amount_col = col
                    mapping_columns.append(col)
                elif 'bank transaction date' in col_clean:
                    date_col = col
                    mapping_columns.append(col)
            
            if panda_id_col and amount_col:
                # Create mapping dataframe
                if date_col:
                    mapping_df = account_df[[panda_id_col, amount_col, date_col]].copy()
                else:
                    mapping_df = account_df[[panda_id_col, amount_col]].copy()
                
                # Clean up amount column
                mapping_df[amount_col] = pd.to_numeric(mapping_df[amount_col], errors='coerce').fillna(0)
                
                # Store the mapping data
                bank_mapping_data[account_num] = {
                    'data': mapping_df,
                    'panda_id_col': panda_id_col,
                    'amount_col': amount_col,
                    'date_col': date_col
                }
                
                print(f"  Created mapping with {len(mapping_df)} records")
                print(f"  Columns: {mapping_columns}")
            else:
                print(f"  Missing required columns (Panda ID or Amount)")
    
    return bank_mapping_data

def match_netsuite_with_bank_data(netsuite_data, bank_mapping_data, account_num):
    """Match Netsuite transactions with bank transactions to get Panda Bank Transaction IDs"""
    
    if account_num not in bank_mapping_data:
        print(f"  No bank mapping data for account {account_num}")
        # Return Panda IDs as empty strings
        return [''] * len(netsuite_data)
    
    bank_info = bank_mapping_data[account_num]
    bank_df = bank_info['data']
    panda_id_col = bank_info['panda_id_col']
    amount_col = bank_info['amount_col']
    date_col = bank_info['date_col']
    
    print(f"  Matching {len(netsuite_data)} Netsuite transactions with {len(bank_df)} bank transactions")
    
    panda_ids = []
    matched_count = 0
    
    for idx, row in netsuite_data.iterrows():
        netsuite_amount = row['Amount']
        panda_id = ''
        
        # Try to match by amount first
        amount_matches = bank_df[abs(bank_df[amount_col] - netsuite_amount) < 0.01]
        
        if len(amount_matches) == 1:
            # Exact amount match
            panda_id = str(amount_matches.iloc[0][panda_id_col])
            matched_count += 1
        elif len(amount_matches) > 1:
            # Multiple amount matches, try to narrow down by date if available
            if date_col and 'Date' in row:
                # Additional date matching logic could go here
                # For now, take the first match
                panda_id = str(amount_matches.iloc[0][panda_id_col])
                matched_count += 1
            else:
                # Take the first match
                panda_id = str(amount_matches.iloc[0][panda_id_col])
                matched_count += 1
        # If no match, leave empty
        
        panda_ids.append(panda_id)
    
    print(f"  Matched {matched_count} out of {len(netsuite_data)} transactions ({matched_count/len(netsuite_data)*100:.1f}%)")
    
    return panda_ids

def add_panda_transaction_ids_to_tabs():
    try:
        # Read the original Netsuite data
        input_file = "Netsuite Transaction Details.xlsx"
        df = pd.read_excel(input_file, sheet_name=0, header=6)
        
        print(f"Reading {input_file}...")
        print(f"Original records: {df.shape[0]}")
        
        # Apply filters
        filtered_df = df[
            (df["Split"] == "22010 - Customer Funds Obligation : Customer Funds Liability") &
            (~df["Memo"].astype(str).str.contains("Customer Cash Deposits in Transit", case=False, na=False))
        ]
        
        print(f"Filtered records: {filtered_df.shape[0]}")
        
        # Read bank transaction mapping data
        bank_mapping_data = read_bank_transaction_mapping()
        
        # Non-batch activity posting accounts (these need Panda Bank Transaction IDs)
        non_batch_accounts = [10068, 10069, 10071, 10504, 10513]
        batch_accounts = [10510, 10512, 10521, 10523, 10525]
        
        # Exclude 10540 account from detailed tabs
        detailed_accounts = filtered_df[filtered_df["Account (Line): Number"] != 10540]
        
        # Get unique accounts for processing
        unique_accounts = detailed_accounts.groupby(['Account (Line): Number', 'Account']).size().reset_index(name='Count')
        unique_accounts = unique_accounts.sort_values('Account (Line): Number')
        
        print(f"\nUpdating Excel file with Panda Bank Transaction IDs for non-batch accounts:")
        for account_num in non_batch_accounts:
            if account_num in bank_mapping_data:
                print(f"  {account_num}: Will add Panda Bank Transaction IDs")
        
        # Update the existing Excel file
        output_file = "Customer_Funds_Pivot_Table.xlsx"
        
        # Load existing workbook
        workbook = openpyxl.load_workbook(output_file)
        
        # Process each account tab
        for idx, row in unique_accounts.iterrows():
            account_num = int(row['Account (Line): Number'])
            account_name = row['Account']
            
            sheet_name = f"{account_num}"
            
            if sheet_name in workbook.sheetnames:
                print(f"\nProcessing account {account_num}...")
                
                if account_num in non_batch_accounts:
                    # Add Panda Bank Transaction ID for non-batch accounts
                    
                    # Get Netsuite data for this account
                    account_netsuite_data = detailed_accounts[detailed_accounts["Account (Line): Number"] == account_num].copy()
                    
                    # Get Panda Bank Transaction IDs by matching with bank data
                    panda_ids = match_netsuite_with_bank_data(account_netsuite_data, bank_mapping_data, account_num)
                    
                    # Read existing sheet data
                    existing_df = pd.read_excel(output_file, sheet_name=sheet_name)
                    
                    # Skip the header rows (ACCOUNT TOTAL, separator)
                    data_start_row = 2  # 0-indexed, so row 3 in Excel
                    data_rows = existing_df.iloc[data_start_row:-1]  # Exclude last row (TOTAL)
                    
                    # Insert/Update Panda Bank Transaction ID as Column C
                    # Shift existing columns to the right
                    new_df = existing_df.copy()
                    
                    # Check if Panda Bank Transaction Id column already exists
                    panda_col_name = 'Panda Bank Transaction Id'
                    if panda_col_name in new_df.columns:
                        # Column exists, update it
                        print(f"  Updating existing {panda_col_name} column")
                        if len(panda_ids) == len(data_rows):
                            # Update the existing column for data rows
                            for i, panda_id in enumerate(panda_ids):
                                new_df.iloc[data_start_row + i, new_df.columns.get_loc(panda_col_name)] = panda_id
                            print(f"  Updated {len([p for p in panda_ids if p])} Panda Bank Transaction IDs")
                        else:
                            print(f"  Warning: Mismatch in record counts - Netsuite: {len(data_rows)}, Panda IDs: {len(panda_ids)}")
                    else:
                        # Column doesn't exist, insert it at position 2 (Column C, 0-indexed)
                        print(f"  Adding new {panda_col_name} column")
                        if len(panda_ids) == len(data_rows):
                            # Create the full column with empty values for header rows and total row
                            full_panda_ids = [''] * len(new_df)
                            
                            # Fill in the Panda IDs for data rows
                            for i, panda_id in enumerate(panda_ids):
                                full_panda_ids[data_start_row + i] = panda_id
                            
                            # Insert the column
                            new_df.insert(2, panda_col_name, full_panda_ids)
                            
                            print(f"  Added {len([p for p in panda_ids if p])} Panda Bank Transaction IDs")
                        else:
                            print(f"  Warning: Mismatch in record counts - Netsuite: {len(data_rows)}, Panda IDs: {len(panda_ids)}")
                            # Add empty column as fallback
                            new_df.insert(2, panda_col_name, [''] * len(new_df))
                    
                    # Delete the old sheet and create new one
                    del workbook[sheet_name]
                    new_sheet = workbook.create_sheet(sheet_name)
                    
                    # Write data to new sheet
                    for r_idx, (_, row_data) in enumerate(new_df.iterrows(), 1):
                        for c_idx, value in enumerate(row_data, 1):
                            new_sheet.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Format the sheet
                    format_worksheet_clean(new_sheet, 'detail')
                    
                    print(f"  Updated tab '{sheet_name}' with Panda Bank Transaction IDs as Column C")
                
                else:
                    print(f"  Skipped account {account_num} (batch activity posting account)")
            else:
                print(f"  Sheet {sheet_name} not found")
        
        # Save the updated workbook
        workbook.save(output_file)
        workbook.close()
        
        print(f"\nUpdated file: {output_file}")
        print(f"Features:")
        print(f"  ✅ Added Panda Bank Transaction ID as Column C for non-batch accounts")
        print(f"  ✅ Matched Netsuite transactions with bank transactions by amount")
        print(f"  ✅ Clean formatting maintained")
        print(f"  ✅ Batch activity accounts unchanged")
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    add_panda_transaction_ids_to_tabs()


#!/usr/bin/env python3

import os
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
import io
import zipfile
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

class BankFeesProcessor:
    def __init__(self):
        self.bank_mapping_data = {}
        self.netsuite_data = None
        self.processed_data = {}
        
        # Account mapping: Netsuite Account -> (Account Number, Bank Transaction Files)
        self.account_mapping = {
            10068: "Chase Reverse Wire Payrolls",
            10069: "Chase Recovery Ops", 
            10071: "Chase Wire In",
            10504: "Chase 3rd Party Processors",
            10513: "PNC Customer Wire Ins"
        }
        
        # Non-batch and batch account classifications
        self.non_batch_accounts = [10068, 10069, 10071, 10504, 10513]
        self.batch_accounts = [10510, 10512, 10521, 10523, 10525]
    
    def format_worksheet_professional(self, worksheet, sheet_type='detail'):
        """Format worksheet with professional financial institution styling"""
        
        # Define professional styles
        header_font = Font(name='Calibri', bold=True, size=11)
        total_font = Font(name='Calibri', bold=True, size=11)
        data_font = Font(name='Calibri', size=10)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Define alternating row colors
        light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format based on sheet type
        if sheet_type == 'detail':
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Format data rows with alternating colors
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # Apply alternating row colors
                    if row_idx % 2 == 0:
                        cell.fill = light_fill
                    
                    # Format by column type
                    header_value = str(worksheet.cell(1, cell.column).value).strip().lower()
                    if 'amount' in header_value:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif 'panda bank transaction id' in header_value or 'document number' in header_value:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif 'date' in header_value:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Bold total rows
                    if str(cell.value).startswith('ACCOUNT ') and 'TOTAL' in str(cell.value):
                        cell.font = total_font
                    elif cell.value == 'TOTAL':
                        cell.font = total_font
        
        elif sheet_type == 'pivot':
            # Format pivot table headers
            for row in worksheet.iter_rows(min_row=3, max_row=3):
                for cell in row:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            
            # Format pivot data
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=4), 4):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # Apply alternating row colors
                    if row_idx % 2 == 0:
                        cell.fill = light_fill
                    
                    # Format amount columns
                    if cell.column in [2, 3, 4]:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Bold subtotal and grand total rows
                    if cell.value in ['Subtotal', 'Grand Total']:
                        cell.font = total_font
    
    def read_bank_transaction_files(self, file_paths):
        """Read and process bank transaction files"""
        
        for account_num, account_name in self.account_mapping.items():
            print(f"\nProcessing account {account_num}: {account_name}")
            
            # Find matching files for this account
            account_files = [f for f in file_paths if str(account_num) in f or 
                           any(str(i) in f for i in [6, 7, 9, 18, 21] if 
                               (account_num == 10068 and i == 6) or
                               (account_num == 10069 and i == 7) or  
                               (account_num == 10071 and i == 9) or
                               (account_num == 10513 and i == 18) or
                               (account_num == 10504 and i == 21))]
            
            if not account_files:
                continue
                
            combined_data = []
            for file_path in account_files:
                try:
                    # Try different configurations
                    configurations = [
                        {'encoding': 'utf-8', 'sep': '\t'},
                        {'encoding': 'utf-8', 'sep': ','},
                        {'encoding': 'utf-16', 'sep': '\t'},
                        {'encoding': 'latin-1', 'sep': '\t'},
                    ]
                    
                    df = None
                    for config in configurations:
                        try:
                            df = pd.read_csv(file_path, encoding=config['encoding'], 
                                           sep=config['sep'], on_bad_lines='skip')
                            if len(df) > 0 and len(df.columns) > 5:
                                break
                        except:
                            continue
                    
                    if df is not None:
                        combined_data.append(df)
                        print(f"  Read {len(df)} records from {os.path.basename(file_path)}")
                        
                except Exception as e:
                    print(f"  Error reading {file_path}: {e}")
            
            if combined_data:
                # Combine all dataframes for this account
                account_df = pd.concat(combined_data, ignore_index=True)
                
                # Find required columns
                panda_id_col = None
                amount_col = None
                date_col = None
                
                for col in account_df.columns:
                    col_clean = str(col).strip().lower()
                    if 'panda bank transaction id' in col_clean:
                        panda_id_col = col
                    elif 'signed amount' in col_clean:
                        amount_col = col
                    elif 'bank transaction date' in col_clean:
                        date_col = col
                
                if panda_id_col and amount_col:
                    mapping_columns = [panda_id_col, amount_col]
                    if date_col:
                        mapping_columns.append(date_col)
                    
                    mapping_df = account_df[mapping_columns].copy()
                    mapping_df[amount_col] = pd.to_numeric(mapping_df[amount_col], errors='coerce').fillna(0)
                    
                    self.bank_mapping_data[account_num] = {
                        'data': mapping_df,
                        'panda_id_col': panda_id_col,
                        'amount_col': amount_col,
                        'date_col': date_col
                    }
                    
                    print(f"  Created mapping with {len(mapping_df)} records")
    
    def load_netsuite_data(self, file_path):
        """Load and filter NetSuite transaction data"""
        
        try:
            df = pd.read_excel(file_path, sheet_name=0, header=6)
            print(f"Loaded {len(df)} NetSuite records")
            
            # Apply filters
            filtered_df = df[
                (df["Split"] == "22010 - Customer Funds Obligation : Customer Funds Liability") &
                (~df["Memo"].astype(str).str.contains("Customer Cash Deposits in Transit", case=False, na=False))
            ]
            
            print(f"Filtered to {len(filtered_df)} records")
            self.netsuite_data = filtered_df
            
            return True
            
        except Exception as e:
            print(f"Error loading NetSuite data: {e}")
            return False
    
    def match_transactions(self, account_num, netsuite_data):
        """Match NetSuite transactions with bank data to get Panda Bank Transaction IDs"""
        
        if account_num not in self.bank_mapping_data:
            return [''] * len(netsuite_data)
        
        bank_info = self.bank_mapping_data[account_num]
        bank_df = bank_info['data']
        panda_id_col = bank_info['panda_id_col']
        amount_col = bank_info['amount_col']
        
        panda_ids = []
        matched_count = 0
        
        for idx, row in netsuite_data.iterrows():
            netsuite_amount = row['Amount']
            panda_id = ''
            
            # Match by amount
            amount_matches = bank_df[abs(bank_df[amount_col] - netsuite_amount) < 0.01]
            
            if len(amount_matches) >= 1:
                panda_id = str(amount_matches.iloc[0][panda_id_col])
                matched_count += 1
            
            panda_ids.append(panda_id)
        
        print(f"  Matched {matched_count}/{len(netsuite_data)} transactions ({matched_count/len(netsuite_data)*100:.1f}%)")
        return panda_ids
    
    def generate_excel_report(self, output_path):
        """Generate comprehensive Excel report with all account tabs"""
        
        if self.netsuite_data is None:
            return False
        
        # Exclude 10540 account from detailed tabs
        detailed_accounts = self.netsuite_data[self.netsuite_data["Account (Line): Number"] != 10540]
        
        # Get unique accounts
        unique_accounts = detailed_accounts.groupby(['Account (Line): Number', 'Account']).size().reset_index(name='Count')
        unique_accounts = unique_accounts.sort_values('Account (Line): Number')
        
        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Process each account
        for idx, row in unique_accounts.iterrows():
            account_num = int(row['Account (Line): Number'])
            account_name = row['Account']
            
            print(f"\nProcessing account {account_num}: {account_name}")
            
            # Get account data
            account_data = detailed_accounts[detailed_accounts["Account (Line): Number"] == account_num].copy()
            
            # Create sheet
            sheet_name = f"{account_num}"
            sheet = workbook.create_sheet(sheet_name)
            
            # Add account total header
            sheet.cell(row=1, column=1, value=f"ACCOUNT {account_num} TOTAL")
            sheet.cell(row=1, column=2, value=account_data['Amount'].sum())
            
            # Add separator row
            sheet.cell(row=2, column=1, value="")
            
            # Prepare columns
            columns = ['Date', 'Document Number']
            
            # Add Panda Bank Transaction ID for non-batch accounts
            if account_num in self.non_batch_accounts:
                columns.append('Panda Bank Transaction Id')
                panda_ids = self.match_transactions(account_num, account_data)
            else:
                panda_ids = []
            
            columns.extend(['Memo', 'Amount'])
            
            # Add headers
            for col_idx, header in enumerate(columns, 1):
                sheet.cell(row=3, column=col_idx, value=header)
            
            # Add data
            for row_idx, (_, data_row) in enumerate(account_data.iterrows(), 4):
                col_idx = 1
                sheet.cell(row=row_idx, column=col_idx, value=data_row['Date'])
                col_idx += 1
                sheet.cell(row=row_idx, column=col_idx, value=data_row['Document Number'])
                col_idx += 1
                
                # Add Panda ID if non-batch account
                if account_num in self.non_batch_accounts and panda_ids:
                    sheet.cell(row=row_idx, column=col_idx, value=panda_ids[row_idx-4] if row_idx-4 < len(panda_ids) else '')
                    col_idx += 1
                
                sheet.cell(row=row_idx, column=col_idx, value=data_row['Memo'])
                col_idx += 1
                sheet.cell(row=row_idx, column=col_idx, value=data_row['Amount'])
            
            # Add total row
            total_row = len(account_data) + 4
            sheet.cell(row=total_row, column=1, value="TOTAL")
            sheet.cell(row=total_row, column=len(columns), value=account_data['Amount'].sum())
            
            # Format sheet
            self.format_worksheet_professional(sheet, 'detail')
        
        # Save workbook
        workbook.save(output_path)
        workbook.close()
        
        print(f"\nGenerated Excel report: {output_path}")
        return True

# Global processor instance
processor = BankFeesProcessor()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file uploads"""
    
    if 'bank_files' not in request.files and 'netsuite_file' not in request.files:
        flash('No files selected', 'error')
        return redirect(request.url)
    
    bank_files = request.files.getlist('bank_files')
    netsuite_file = request.files.get('netsuite_file')
    
    # Save uploaded files
    bank_file_paths = []
    netsuite_file_path = None
    
    try:
        # Save bank transaction files
        for file in bank_files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                bank_file_paths.append(file_path)
        
        # Save NetSuite file
        if netsuite_file and netsuite_file.filename:
            filename = secure_filename(netsuite_file.filename)
            netsuite_file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            netsuite_file.save(netsuite_file_path)
        
        # Process files
        if bank_file_paths:
            processor.read_bank_transaction_files(bank_file_paths)
        
        if netsuite_file_path:
            processor.load_netsuite_data(netsuite_file_path)
        
        flash('Files uploaded and processed successfully', 'success')
        return jsonify({'status': 'success', 'message': 'Files processed successfully'})
        
    except Exception as e:
        flash(f'Error processing files: {str(e)}', 'error')
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/generate_report')
def generate_report():
    """Generate Excel report"""
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Customer_Funds_Analysis_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        
        success = processor.generate_excel_report(output_path)
        
        if success:
            return send_file(output_path, as_attachment=True, download_name=output_filename)
        else:
            flash('Error generating report', 'error')
            return redirect(url_for('index'))
            
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/status')
def status():
    """Get processing status"""
    
    status_info = {
        'bank_accounts_loaded': len(processor.bank_mapping_data),
        'netsuite_loaded': processor.netsuite_data is not None,
        'netsuite_records': len(processor.netsuite_data) if processor.netsuite_data is not None else 0,
        'accounts': list(processor.account_mapping.keys()),
        'loaded_accounts': list(processor.bank_mapping_data.keys())
    }
    
    return jsonify(status_info)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)


